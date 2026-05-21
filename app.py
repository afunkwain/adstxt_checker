"""
AdsTxt Checker — Python Backend
================================
Flask server that checks domains for google.com in their ads.txt files.

HOW TO RUN:
  1. pip install flask flask-cors requests openpyxl
  2. python app.py
  3. Open http://localhost:5000 in your browser
"""

import asyncio
import csv
import io
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import openpyxl
from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

# ── Config ─────────────────────────────────────────────────────────────────────
TIMEOUT        = 8          # seconds per request
MAX_WORKERS    = 30         # concurrent threads
GOOGLE_PATTERN = re.compile(r'google\.com', re.IGNORECASE)

# In-memory job store  {job_id: {...}}
jobs = {}


# ── Helpers ────────────────────────────────────────────────────────────────────

def clean_domain(raw: str) -> str:
    """Strip protocol, paths, whitespace — keep bare domain."""
    raw = raw.strip()
    raw = re.sub(r'^https?://', '', raw, flags=re.IGNORECASE)
    raw = raw.split('/')[0]          # drop any path
    raw = raw.split('?')[0]          # drop query string
    return raw.lower()


def dedupe_domains(domains: list) -> list:
    """Deduplicate domains while preserving original order."""
    seen = set()
    unique = []
    for d in domains:
        if d and d not in seen:
            seen.add(d)
            unique.append(d)
    return unique


def check_domain(domain: str) -> dict:
    """Fetch /ads.txt for a domain and look for google.com."""
    result = {
        "domain":      domain,
        "status":      "Failed",
        "google":      "No",
        "status_code": None,
        "error":       "",
    }

    for scheme in ("https", "http"):
        url = f"{scheme}://{domain}/ads.txt"
        try:
            resp = requests.get(
                url,
                timeout=TIMEOUT,
                headers={"User-Agent": "AdsTxtChecker/1.0"},
                allow_redirects=True,
            )
            result["status_code"] = resp.status_code

            if resp.status_code == 404:
                result["status"] = "Not Found"
                return result

            if resp.status_code == 200:
                result["status"] = "Success"
                if GOOGLE_PATTERN.search(resp.text):
                    result["google"] = "Yes"
                return result

            # other 4xx/5xx — try next scheme
            result["status"] = f"HTTP {resp.status_code}"

        except requests.exceptions.SSLError:
            continue                  # fall back to http
        except requests.exceptions.ConnectionError as e:
            result["error"] = "Connection error"
        except requests.exceptions.Timeout:
            result["error"] = "Timeout"
        except Exception as e:
            result["error"] = str(e)[:80]

    if result["status"] == "Failed" and not result["error"]:
        result["error"] = "Could not connect"

    return result


def run_job(job_id: str, domains: list):
    """Background worker — processes all domains and updates job state."""
    job = jobs[job_id]
    job["total"]   = len(domains)
    job["done"]    = 0
    job["results"] = []
    job["started"] = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        future_map = {pool.submit(check_domain, d): d for d in domains}
        for future in as_completed(future_map):
            res = future.result()
            job["results"].append(res)
            job["done"] += 1

    job["finished"] = True
    job["elapsed"]  = round(time.time() - job["started"], 1)


def parse_domains_from_excel(file_bytes: bytes) -> dict:
    """Extract domains and return input summary for an xlsx/xls file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    raw_values = []
    normalized_domains = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and str(cell).strip():
                    raw = str(cell).strip()
                    raw_values.append(raw)
                    cleaned = clean_domain(raw)
                    if cleaned:
                        normalized_domains.append(cleaned)

    unique = dedupe_domains(normalized_domains)
    return {
        "domains": unique,
        "submitted": len(raw_values),
        "normalized": len(normalized_domains),
        "unique": len(unique),
    }


def parse_domains_from_text(text: str) -> dict:
    lines = text.splitlines()
    raw_values = []
    normalized_domains = []

    for line in lines:
        if not line.strip():
            continue
        raw = line.strip()
        raw_values.append(raw)
        cleaned = clean_domain(raw)
        if cleaned:
            normalized_domains.append(cleaned)

    unique = dedupe_domains(normalized_domains)
    return {
        "domains": unique,
        "submitted": len(raw_values),
        "normalized": len(normalized_domains),
        "unique": len(unique),
    }


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


def _extract_domains_from_request():
    """Parse domains from uploaded Excel or JSON body. Returns (domains, counts) or error."""
    submitted_count = 0
    normalized_count = 0

    if "file" in request.files:
        f = request.files["file"]
        try:
            parsed = parse_domains_from_excel(f.read())
        except Exception as e:
            return None, (f"Could not read Excel file: {e}", 400)
        domains = parsed["domains"]
        submitted_count = parsed["submitted"]
        normalized_count = parsed["normalized"]

    elif request.is_json:
        data = request.get_json() or {}
        raw = data.get("domains", "")
        parsed = parse_domains_from_text(raw)
        domains = parsed["domains"]
        submitted_count = parsed["submitted"]
        normalized_count = parsed["normalized"]

    else:
        return None, ("Upload a file or send domains as JSON.", 400)

    if not domains:
        return None, ("No valid domains found.", 400)

    if len(domains) > 10_000:
        return None, ("Max 10 000 domains per run.", 400)

    unique_count = len(domains)
    duplicates_removed = max(normalized_count - unique_count, 0)
    invalid_skipped = max(submitted_count - normalized_count, 0)

    return {
        "domains": domains,
        "submitted": submitted_count,
        "unique": unique_count,
        "duplicates_removed": duplicates_removed,
        "invalid_skipped": invalid_skipped,
    }, None


@app.route("/api/parse", methods=["POST"])
def parse_only():
    """Parse domains from upload or text without starting a scan."""
    payload, err = _extract_domains_from_request()
    if err:
        return jsonify({"error": err[0]}), err[1]
    return jsonify(payload)


@app.route("/api/check_batch", methods=["POST"])
def check_batch():
    """Check a batch of domains synchronously (works on serverless)."""
    if not request.is_json:
        return jsonify({"error": "JSON body required."}), 400

    data = request.get_json() or {}
    domains = data.get("domains") or []
    if not isinstance(domains, list):
        return jsonify({"error": "domains must be a list."}), 400

    domains = [clean_domain(str(d)) for d in domains if str(d).strip()]
    domains = dedupe_domains(domains)

    if not domains:
        return jsonify({"error": "No valid domains in batch."}), 400

    if len(domains) > 100:
        return jsonify({"error": "Max 100 domains per batch."}), 400

    try:
        workers = int(data.get("workers", MAX_WORKERS))
    except (TypeError, ValueError):
        workers = MAX_WORKERS
    workers = min(max(workers, 1), 50)

    results = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_map = {pool.submit(check_domain, d): d for d in domains}
        for future in as_completed(future_map):
            results.append(future.result())

    results.sort(key=lambda r: r["domain"])
    return jsonify({"results": results})


@app.route("/api/check", methods=["POST"])
def start_check():
    payload, err = _extract_domains_from_request()
    if err:
        return jsonify({"error": err[0]}), err[1]

    domains = payload["domains"]
    submitted_count = payload["submitted"]
    unique_count = payload["unique"]
    duplicates_removed = payload["duplicates_removed"]
    invalid_skipped = payload["invalid_skipped"]

    job_id = f"job_{int(time.time()*1000)}"
    jobs[job_id] = {
        "finished": False,
        "total": 0,
        "done": 0,
        "results": [],
        "summary": {
            "job_id": job_id,
            "run_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "submitted": submitted_count,
            "unique": unique_count,
            "duplicates_removed": duplicates_removed,
            "invalid_skipped": invalid_skipped,
        },
    }

    thread = threading.Thread(target=run_job, args=(job_id, domains), daemon=True)
    thread.start()

    return jsonify({
        "job_id": job_id,
        "total": unique_count,
        "submitted": submitted_count,
        "unique": unique_count,
        "duplicates_removed": duplicates_removed,
        "invalid_skipped": invalid_skipped,
    })


@app.route("/api/status/<job_id>")
def job_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "finished": job["finished"],
        "total":    job["total"],
        "done":     job["done"],
        "results":  job["results"],
        "elapsed":  job.get("elapsed"),
    })


def _build_csv_bytes(results: list, summary: dict) -> bytes:
    output = io.StringIO()
    output.write("Run Summary\n")
    output.write(f"Job ID,{summary.get('job_id', '')}\n")
    output.write(f"Run Timestamp,{summary.get('run_timestamp', '')}\n")
    output.write(f"Submitted,{summary.get('submitted', '')}\n")
    output.write(f"Unique Scanned,{summary.get('unique', '')}\n")
    output.write(f"Duplicates Removed,{summary.get('duplicates_removed', '')}\n")
    output.write(f"Invalid Skipped,{summary.get('invalid_skipped', '')}\n\n")

    writer = csv.DictWriter(
        output,
        fieldnames=["domain", "status", "google", "status_code", "error"],
        extrasaction="ignore",
    )
    writer.writeheader()
    writer.writerows(results)
    output.seek(0)
    return output.getvalue().encode()


def _build_excel_bytes(results: list, summary: dict) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AdsTxt Results"

    ws.append(["Run Summary", "Value"])
    ws.append(["Job ID", summary.get("job_id", "")])
    ws.append(["Run Timestamp", summary.get("run_timestamp", "")])
    ws.append(["Submitted", summary.get("submitted", "")])
    ws.append(["Unique Scanned", summary.get("unique", "")])
    ws.append(["Duplicates Removed", summary.get("duplicates_removed", "")])
    ws.append(["Invalid Skipped", summary.get("invalid_skipped", "")])
    ws.append([])

    headers = ["Domain", "Status", "Google.com Present", "HTTP Code", "Error"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="00E5A0")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws["A1"].alignment = Alignment(horizontal="left")
    ws["B1"].alignment = Alignment(horizontal="center")

    table_header_row = 9
    for cell in ws[table_header_row]:
        cell.font = Font(bold=True, color="00E5A0")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    green_fill = PatternFill("solid", fgColor="D4EDDA")
    red_fill = PatternFill("solid", fgColor="F8D7DA")

    for row in results:
        ws.append([
            row["domain"],
            row["status"],
            row["google"],
            row["status_code"] or "",
            row["error"],
        ])
        last_row = ws.max_row
        if row["google"] == "Yes":
            ws.cell(last_row, 3).fill = green_fill
        else:
            ws.cell(last_row, 3).fill = red_fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.route("/api/export", methods=["POST"])
def export_results_post():
    """Export scan results sent in the request body (stateless)."""
    if not request.is_json:
        return jsonify({"error": "JSON body required."}), 400

    data = request.get_json() or {}
    results = data.get("results") or []
    summary = data.get("summary") or {}
    fmt = (data.get("format") or "csv").lower()

    if not results:
        return jsonify({"error": "No results to export."}), 400

    if fmt == "excel":
        return send_file(
            io.BytesIO(_build_excel_bytes(results, summary)),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="adstxt_results.xlsx",
        )

    return send_file(
        io.BytesIO(_build_csv_bytes(results, summary)),
        mimetype="text/csv",
        as_attachment=True,
        download_name="adstxt_results.csv",
    )


@app.route("/api/export/<job_id>")
def export_csv(job_id):
    job = jobs.get(job_id)
    if not job or not job["finished"]:
        return jsonify({"error": "Job not ready"}), 404

    summary = job.get("summary", {})

    return send_file(
        io.BytesIO(_build_csv_bytes(job["results"], summary)),
        mimetype="text/csv",
        as_attachment=True,
        download_name="adstxt_results.csv",
    )


@app.route("/api/export_excel/<job_id>")
def export_excel(job_id):
    job = jobs.get(job_id)
    if not job or not job["finished"]:
        return jsonify({"error": "Job not ready"}), 404

    summary = job.get("summary", {})
    return send_file(
        io.BytesIO(_build_excel_bytes(job["results"], summary)),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="adstxt_results.xlsx",
    )


if __name__ == "__main__":
    print("\n🚀  AdsTxt Checker running at http://localhost:5000\n")
    app.run(debug=True, port=5000)
